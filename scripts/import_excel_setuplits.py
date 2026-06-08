#!/usr/bin/env python3
"""
Excel On Air List データを setuplits.html の SETUPLIST_DATA に取り込む

Usage:
    python scripts/import_excel_setuplits.py --excel "path/to/file.xlsx"
"""

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

_START = "// <<SETUPLIST_DATA_START>>"
_END   = "// <<SETUPLIST_DATA_END>>"
_HTML  = Path(__file__).parent.parent / "setuplits.html"


def parse_sheet_date(sheet_name: str) -> str:
    """'2021.4.13' → '2021-04-13'"""
    parts = sheet_name.strip().split(".")
    if len(parts) == 3:
        try:
            y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            return f"{y:04d}-{m:02d}-{d:02d}"
        except ValueError:
            pass
    return sheet_name


def parse_episode_no(title: str):
    """タイトルから '#29' の数字を抽出"""
    m = re.search(r"#(\d+)", str(title))
    return int(m.group(1)) if m else None


def find_biko_col(sheet):
    """備考列のインデックス（1始まり）とデータ開始行を返す"""
    for row_idx in range(2, 8):
        for col_idx in range(1, 10):
            val = str(sheet.cell(row=row_idx, column=col_idx).value or "").strip()
            if "備考" in val:
                return col_idx, row_idx + 1
    # デフォルト: D列(4)、3行目からデータ
    return 4, 3


def extract_special_feature(sheet, biko_col: int, data_start: int):
    """備考列から特集名を抽出（最頻値、'特集'含む値を優先）"""
    values = []
    for row_idx in range(data_start, sheet.max_row + 1):
        val = sheet.cell(row=row_idx, column=biko_col).value
        if val:
            v = str(val).strip()
            if v:
                values.append(v)

    if not values:
        return None

    tokushu = [v for v in values if "特集" in v]
    if tokushu:
        return Counter(tokushu).most_common(1)[0][0]

    return Counter(values).most_common(1)[0][0]


def extract_tracks(sheet, biko_col: int, data_start: int):
    """全曲データを抽出して tracks リストを返す"""
    tracks = []
    for row_idx in range(data_start, sheet.max_row + 1):
        no_val  = sheet.cell(row=row_idx, column=1).value
        title   = sheet.cell(row=row_idx, column=2).value
        artist  = sheet.cell(row=row_idx, column=3).value
        note_v  = sheet.cell(row=row_idx, column=biko_col).value

        # No. 列が数値でない行はスキップ（フッター・空行など）
        if not isinstance(no_val, (int, float)):
            continue

        title_s  = str(title).strip()  if title  else ""
        artist_s = str(artist).strip() if artist else ""
        note_s   = str(note_v).strip() if note_v else ""

        # note の正規化：洋楽枠優先、次に備考あり→特集枠
        if "洋楽" in artist_s:
            note = "洋楽枠"
        elif note_s:
            note = "特集枠"
        else:
            note = None

        tracks.append({
            "no":     int(no_val),
            "title":  title_s,
            "artist": artist_s,
            "note":   note,
        })

    return tracks if tracks else None


def read_html_data(html_path: Path) -> list:
    """HTML から現在の SETUPLIST_DATA を読み取る"""
    html = html_path.read_text(encoding="utf-8")
    m = re.search(
        re.escape(_START) + r".*?const SETUPLIST_DATA = (\[.*?\]);" + r".*?" + re.escape(_END),
        html,
        re.DOTALL,
    )
    if not m:
        raise ValueError(f"SETUPLIST_DATA マーカーが見つかりません: {html_path}")
    return json.loads(m.group(1))


def write_html_data(html_path: Path, data: list) -> None:
    """更新した SETUPLIST_DATA を HTML に書き戻す"""
    html = html_path.read_text(encoding="utf-8")
    json_str = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    new_block = f"{_START}\n  const SETUPLIST_DATA = {json_str};\n  {_END}"

    pattern = re.compile(
        re.escape(_START) + r".*?" + re.escape(_END),
        re.DOTALL,
    )
    html_new = pattern.sub(new_block, html)
    html_path.write_text(html_new, encoding="utf-8")


def process_excel(excel_path: Path, existing_data: list):
    """Excel を処理して (更新済みデータ, 追加数, 更新数, スキップ数) を返す"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    stem = excel_path.stem
    try:
        puy_year = int(stem[:4])
    except ValueError:
        print(f"Warning: ファイル名から puy_year を取得できません ({stem})", file=sys.stderr)
        puy_year = None

    # episode_no → レコードのマップ（tracks更新のため）
    existing_map = {rec["episode_no"]: rec for rec in existing_data}
    existing_nos = set(existing_map.keys())

    new_records = []
    added   = 0
    updated = 0
    skipped = 0

    for sheet_name in wb.sheetnames:
        air_date = parse_sheet_date(sheet_name)
        sheet = wb[sheet_name]

        # 先頭5行のすべてのセルから episode_no を探す
        episode_no = None
        for row_idx in range(1, 6):
            for cell in sheet[row_idx]:
                if cell.value:
                    no = parse_episode_no(str(cell.value).strip())
                    if no is not None:
                        episode_no = no
                        break
            if episode_no is not None:
                break

        if episode_no is None:
            print(f"  [SKIP] シート '{sheet_name}': episode_no が見つかりません")
            skipped += 1
            continue

        biko_col, data_start = find_biko_col(sheet)
        tracks = extract_tracks(sheet, biko_col, data_start)

        if episode_no in existing_nos:
            existing_rec = existing_map[episode_no]
            # tracks がまだ未登録の場合のみ更新
            if not existing_rec.get("tracks"):
                existing_rec["tracks"] = tracks
                updated += 1
                print(f"  [UPD ] episode_no={episode_no}: tracks {len(tracks) if tracks else 0}曲 追加")
            else:
                print(f"  [SKIP] episode_no={episode_no}: 既存（tracks登録済み）")
                skipped += 1
            continue

        special_feature = extract_special_feature(sheet, biko_col, data_start)

        record = {
            "episode_no":    episode_no,
            "puy_year":      puy_year,
            "air_date":      air_date,
            "special_feature": special_feature,
            "image_file":    None,
            "tracks":        tracks,
        }
        new_records.append(record)
        existing_nos.add(episode_no)
        added += 1
        print(f"  [ADD ] episode_no={episode_no}, air_date={air_date}, "
              f"puy_year={puy_year}, special_feature={special_feature}, "
              f"tracks={len(tracks) if tracks else 0}曲")

    updated_list = sorted(existing_data + new_records, key=lambda x: x["episode_no"])
    return updated_list, added, updated, skipped


def main():
    parser = argparse.ArgumentParser(description="Excel On Air List を SETUPLIST_DATA に取り込む")
    parser.add_argument("--excel", required=True, help="Excel ファイルのパス")
    parser.add_argument("--html", default=str(_HTML), help="setuplits.html のパス")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    html_path  = Path(args.html)

    if not excel_path.exists():
        print(f"Error: Excel ファイルが見つかりません: {excel_path}", file=sys.stderr)
        sys.exit(1)
    if not html_path.exists():
        print(f"Error: HTML ファイルが見つかりません: {html_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Excel : {excel_path}")
    print(f"HTML  : {html_path}")
    print()

    existing_data = read_html_data(html_path)
    print(f"既存レコード数: {len(existing_data)}")
    print()

    updated_data, added, updated, skipped = process_excel(excel_path, existing_data)

    if added > 0 or updated > 0:
        write_html_data(html_path, updated_data)
        print(f"\n{html_path} を更新しました (合計 {len(updated_data)} 件)")

    print(f"\n==============================")
    print(f"新規追加: {added} 件")
    print(f"tracks更新: {updated} 件")
    print(f"スキップ: {skipped} 件")
    print(f"==============================")


if __name__ == "__main__":
    main()
